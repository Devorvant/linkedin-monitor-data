from __future__ import annotations

import argparse
import base64
import csv
import os
import ctypes
import json
import re
import subprocess
import sys
import time
from datetime import date, datetime, time as dt_time, timedelta
from email import policy
from email.parser import BytesParser
from pathlib import Path
from urllib.parse import urlparse, parse_qs
from urllib.error import HTTPError
from urllib.request import Request, urlopen

try:
    import pyautogui
except ImportError:
    pyautogui = None

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None


_SCRIPT_DIR = Path(__file__).resolve().parent
# Если worker скачан диспетчером в папку cache, рабочий корень проекта
# находится уровнем выше. При обычном ручном запуске ROOT остаётся
# папкой самого файла.
ROOT = _SCRIPT_DIR.parent if _SCRIPT_DIR.name.lower() == "cache" else _SCRIPT_DIR

CHROME_PATHS = [
    Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
    Path(r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
]
URLS = {
    "feed": "https://www.linkedin.com/feed/",
    "jobs": "https://www.linkedin.com/jobs/",
    "notifications": "https://www.linkedin.com/notifications/",
}


# ----------------------------------------------------------------------
# GUI-профили для двух компьютеров
# ----------------------------------------------------------------------
DEVICE_PRESETS = {
    "1": {
        "name": "Ноутбук (старые рабочие координаты)",
        "scroll_x": 0.82,
        "scroll_y": 0.55,
        "save_right_x": 0.88,
        "save_right_y": 0.46,
        "save_menu_x": 0.74,
        "save_menu_y_offset": 0.155,
        "close_tab_x": 0.343,
        "close_tab_y": 0.022,
    },
    "2": {
        "name": "Компьютер 1280x1024",
        "scroll_x": 0.938,
        "scroll_y": 0.547,
        "save_right_x": 0.938,
        "save_right_y": 0.488,

        # Сохраняем старую логику контекстного меню:
        # пункт "Сохранить как..." выбирается по относительной позиции.
        "save_menu_x": 0.798,
        # На новом компьютере 0.155 попадало на следующую строку "Печать".
        # Поднимаем клик примерно на одну строку меню выше.
        "save_menu_y_offset": 0.130,

        # На компьютере 2 стартуем без заранее открытой LinkedIn-вкладки.
        # Поэтому открытая автоматизацией страница является первой/единственной.
        # Для 1280x1024 крестик вкладки находится примерно около (256, 20).
        "close_tab_x": 0.200,
        "close_tab_y": 0.020,
    },
}

ACTIVE_DEVICE = DEVICE_PRESETS["1"]


def choose_device() -> None:
    global ACTIVE_DEVICE

    print("\\nНа каком компьютере работаем?")
    print("  1 - Ноутбук")
    print("  2 - Компьютер 1280x1024")

    while True:
        choice = input("Выберите [1-2]: ").strip()
        if choice in DEVICE_PRESETS:
            ACTIVE_DEVICE = DEVICE_PRESETS[choice]
            print(f"Выбран GUI-профиль: {ACTIVE_DEVICE['name']}\\n")
            return

        print("Введите 1 или 2.")



# GitHub publication settings
REPORTS_DIR = ROOT / "reports"
LATEST_JSON = REPORTS_DIR / "latest.json"

GITHUB_OWNER = "Devorvant"
GITHUB_REPO = "linkedin-monitor-data"
GITHUB_BRANCH = "main"
GITHUB_PATH = "latest.json"

TOKEN_FILES = [
    ROOT / "github_token.txt",
    REPORTS_DIR / "github_token.txt",
]


def load_config() -> dict:
    path = ROOT / "config.json"
    return json.loads(path.read_text(encoding="utf-8"))


def find_chrome() -> Path:
    for path in CHROME_PATHS:
        if path.exists():
            return path
    raise FileNotFoundError("Google Chrome не найден в стандартных папках.")


def focus_linkedin_chrome(maximize: bool = False) -> bool:
    """Находит видимое окно LinkedIn в Chrome и делает его активным."""
    if sys.platform != "win32":
        return False

    user32 = ctypes.windll.user32
    found = []

    WNDENUMPROC = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p)

    @WNDENUMPROC
    def enum_proc(hwnd, lparam):
        if not user32.IsWindowVisible(hwnd):
            return True

        length = user32.GetWindowTextLengthW(hwnd)
        if length <= 0:
            return True

        buf = ctypes.create_unicode_buffer(length + 1)
        user32.GetWindowTextW(hwnd, buf, length + 1)
        title = buf.value.lower()

        if "linkedin" in title and "chrome" in title:
            found.append(hwnd)

        return True

    user32.EnumWindows(enum_proc, 0)

    if not found:
        print("Не удалось найти окно LinkedIn/Chrome.")
        return False

    hwnd = found[-1]

    SW_RESTORE = 9
    SW_MAXIMIZE = 3

    user32.ShowWindow(hwnd, SW_RESTORE)
    time.sleep(0.3)

    if maximize:
        user32.ShowWindow(hwnd, SW_MAXIMIZE)
        time.sleep(0.5)

    user32.BringWindowToTop(hwnd)
    user32.SetForegroundWindow(hwnd)
    time.sleep(0.8)

    return True


def activate_and_open(url: str, wait_seconds: float) -> None:
    chrome = find_chrome()
    subprocess.Popen([str(chrome), url])
    print(f"Открываю обычный Chrome: {url}")
    print(f"Ожидание загрузки: {wait_seconds:.0f} с")
    time.sleep(wait_seconds)


def click_show_more(config: dict) -> int:
    """Опционально ищет на экране шаблоны кнопок и кликает по ним."""
    if pyautogui is None or not config.get("click_show_more", False):
        return 0

    template_names = config.get("show_more_templates", [])
    confidence = float(config.get("image_confidence", 0.82))
    clicked = 0

    for name in template_names:
        template = ROOT / "templates" / name
        if not template.exists():
            continue
        try:
            boxes = list(pyautogui.locateAllOnScreen(str(template), confidence=confidence))
        except Exception as exc:
            print(f"Поиск шаблона {name} пропущен: {exc}")
            continue

        for box in boxes[:20]:
            try:
                pyautogui.click(pyautogui.center(box))
                clicked += 1
                time.sleep(0.4)
            except Exception:
                pass
    return clicked


def scroll_page(config: dict) -> None:
    if pyautogui is None:
        raise RuntimeError("Не установлен pyautogui.")

    count = int(config.get("scroll_count", 8))
    delay = float(config.get("scroll_delay_seconds", 2.5))
    method = config.get("scroll_method", "pagedown")

    # Клик ближе к правому краю окна, чтобы фокус получила страница.
    # По вертикали держимся в середине, чтобы не попасть в кнопки Chrome.
    width, height = pyautogui.size()
    x = int(width * ACTIVE_DEVICE["scroll_x"])
    y = int(height * ACTIVE_DEVICE["scroll_y"])
    pyautogui.moveTo(x, y, duration=0.4)
    pyautogui.click()
    time.sleep(0.5)

    for index in range(count):
        clicked = click_show_more(config)
        if clicked:
            print(f"Раскрыто кнопок: {clicked}")

        if method == "wheel":
            pyautogui.scroll(-int(config.get("wheel_amount", 60)))
        else:
            pyautogui.press("pagedown")

        print(f"Прокрутка {index + 1}/{count}")
        time.sleep(delay)

    clicked = click_show_more(config)
    if clicked:
        print(f"Дополнительно раскрыто кнопок: {clicked}")


def set_windows_clipboard(text: str) -> None:
    """Надёжно кладёт текст в буфер обмена Windows через clip.exe."""
    if sys.platform != "win32":
        raise RuntimeError("Буфер обмена поддерживается только в Windows.")

    result = subprocess.run(
        ["cmd", "/c", "clip"],
        input=text,
        text=True,
        encoding="utf-8",
        capture_output=True,
    )
    if result.returncode != 0:
        raise RuntimeError(f"Не удалось записать в буфер обмена: {result.stderr}")



def snapshot_saved_pages(folder: Path) -> dict[str, tuple[float, int]]:
    """Снимок MHTML/HTML файлов: mtime + size."""
    result = {}
    if not folder.exists():
        return result

    for path in folder.iterdir():
        if not path.is_file():
            continue
        if path.suffix.lower() not in {".mhtml", ".mht", ".html", ".htm"}:
            continue
        try:
            st = path.stat()
            result[str(path.resolve())] = (st.st_mtime, st.st_size)
        except OSError:
            pass
    return result


def wait_for_saved_page_change(
    folder: Path,
    before: dict[str, tuple[float, int]],
    timeout: float = 18.0,
) -> Path | None:
    """Ждёт новый либо изменившийся файл после Save As."""
    deadline = time.time() + timeout
    candidate = None

    while time.time() < deadline:
        changed = []

        for path in folder.iterdir():
            if not path.is_file():
                continue
            if path.suffix.lower() not in {".mhtml", ".mht", ".html", ".htm"}:
                continue

            try:
                st = path.stat()
            except OSError:
                continue

            key = str(path.resolve())
            old = before.get(key)

            if old is None or old != (st.st_mtime, st.st_size):
                changed.append((st.st_mtime, st.st_size, path))

        if changed:
            changed.sort(reverse=True, key=lambda x: (x[0], x[1]))
            newest = changed[0][2]

            # Ждём, пока размер перестанет меняться.
            try:
                size1 = newest.stat().st_size
                time.sleep(0.8)
                size2 = newest.stat().st_size
                if size1 == size2 and size2 > 0:
                    return newest
                candidate = newest
            except OSError:
                pass

        time.sleep(0.5)

    return candidate


def rename_saved_page_for_target(path: Path, target: str) -> Path:
    """Переименовывает фактически сохранённую страницу по типу запуска."""
    if not path.exists():
        return path

    stamp = datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d_%H-%M-%S")
    suffix = path.suffix.lower() or ".mhtml"
    new_path = path.parent / f"linkedin_{target}_{stamp}{suffix}"

    counter = 2
    while new_path.exists() and new_path.resolve() != path.resolve():
        new_path = path.parent / f"linkedin_{target}_{stamp}_{counter}{suffix}"
        counter += 1

    if new_path.resolve() == path.resolve():
        return path

    path.rename(new_path)
    print(f"Переименован файл текущего запуска: {path.name} -> {new_path.name}")
    return new_path


def parse_single_file(path: Path) -> None:
    """Создаёт отчёт только по одному MHTML/HTML файлу."""
    if not path.exists():
        print(f"Файл для парсинга не найден: {path}")
        return

    print(f"Парсинг только текущего файла: {path.name}")

    try:
        html = read_html_or_mhtml(path)

        low_name = path.name.lower()
        if "notification" in low_name or "уведом" in low_name:
            items = extract_notification_records(html, path.name)
            print(f"  Notifications: найдено карточек {len(items)}")
        elif "linkedin_jobs_" in low_name or "ваканс" in low_name or "jobs _ linkedin" in low_name:
            items = extract_jobs_records(html, path.name)
            print(f"  Jobs: найдено вакансий {len(items)}")
        else:
            items = extract_records(html, path.name)
    except Exception as exc:
        items = [{
            "type": "error",
            "title": str(exc),
            "company": "",
            "location": "",
            "text": "",
            "url": "",
            "source_file": path.name,
        }]

    deduped = []
    seen = set()
    for item in items:
        key = (
            item.get("url", "").split("?")[0],
            item.get("title", "")[:300],
            item.get("text", "")[:300],
        )
        if key in seen:
            continue
        seen.add(key)
        item["collected_at"] = datetime.now().isoformat(timespec="seconds")
        deduped.append(item)

    deduped.sort(key=lambda x: (x.get("type", ""), x.get("title", "")))

    reports = ROOT / "reports"
    reports.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    target = "page"
    low = path.name.lower()
    for name in ("feed", "jobs", "notifications"):
        if f"linkedin_{name}_" in low:
            target = name
            break

    base = reports / f"linkedin_{target}_report_{stamp}"

    base.with_suffix(".json").write_text(
        json.dumps(deduped, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    columns = [
        "type", "job_id", "title", "company", "location",
        "url", "source_file", "collected_at", "text"
    ]
    with base.with_suffix(".csv").open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(deduped)

    if pd is not None:
        df = pd.DataFrame(deduped)
        with pd.ExcelWriter(base.with_suffix(".xlsx"), engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="All", index=False)
            if not df.empty:
                for kind in ("job", "post", "notification", "company", "person", "text", "error"):
                    part = df[df["type"] == kind]
                    if not part.empty:
                        part.to_excel(writer, sheet_name=kind, index=False)
        print(f"Excel: {base.with_suffix('.xlsx')}")

    print(f"JSON: {base.with_suffix('.json')}")
    print(f"CSV: {base.with_suffix('.csv')}")
    print(f"Записей текущего файла: {len(deduped)}")


def save_page(target: str, config: dict) -> Path:
    if pyautogui is None:
        raise RuntimeError("Не установлен pyautogui.")

    saved_dir = ROOT / "saved_pages"
    saved_dir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output = saved_dir / f"linkedin_{target}_{stamp}.mhtml"

    print("Перед сохранением повторно активирую окно LinkedIn/Chrome...")
    focus_linkedin_chrome(maximize=True)
    time.sleep(0.8)

    # Сохранение через контекстное меню страницы.
    # Правый клик делаем в пустой области справа, чтобы не попасть по ссылке/картинке.
    width, height = pyautogui.size()
    click_x = int(width * ACTIVE_DEVICE["save_right_x"])
    click_y = int(height * ACTIVE_DEVICE["save_right_y"])

    print(f"Открываю контекстное меню: x={click_x}, y={click_y}")
    pyautogui.moveTo(click_x, click_y, duration=0.35)
    pyautogui.rightClick()
    time.sleep(0.8)

    # В обычном контекстном меню Chrome пункт "Сохранить как..."
    # находится примерно на 4-й строке. Кликаем мышью относительно точки вызова.
    save_x = int(width * ACTIVE_DEVICE["save_menu_x"])
    save_y = click_y + int(height * ACTIVE_DEVICE["save_menu_y_offset"])

    print(f"Нажимаю 'Сохранить как...': x={save_x}, y={save_y}")
    pyautogui.moveTo(save_x, save_y, duration=0.3)
    pyautogui.click()

    time.sleep(float(config.get("save_dialog_wait_seconds", 0.7)))

    # Имя файла не меняем: используем то, которое предлагает Chrome.
    # После короткой паузы сразу подтверждаем сохранение клавишей Enter.
    print("Подтверждаю сохранение Enter (имя Chrome оставляем без изменений)...")
    pyautogui.press("enter")
    time.sleep(float(config.get("save_finish_wait_seconds", 5)))

    # Если Chrome спросил о замене файла, подтверждаем.
    pyautogui.press("left")
    pyautogui.press("enter")
    time.sleep(1)

    print(f"Ожидаемый файл: {output}")
    print(
        "Важно: Chrome должен использовать тип «Веб-страница, один файл (*.mhtml)». "
        "При первом запуске этот тип, возможно, нужно выбрать вручную."
    )
    return output


def automate(target: str) -> None:
    config = load_config()
    url = URLS.get(target, target)

    print(f"GUI-профиль: {ACTIVE_DEVICE['name']}")
    print(
        "\nЧерез 5 секунд начнётся управление клавиатурой и мышью.\n"
        "Не переключайте окна до завершения сохранения.\n"
        "Для аварийной остановки резко переместите мышь в левый верхний угол.\n"
    )
    time.sleep(5)

    pyautogui.FAILSAFE = True
    pyautogui.PAUSE = 0.1

    activate_and_open(url, float(config.get("page_load_wait_seconds", 8)))

    print("Разворачиваю LinkedIn/Chrome на весь экран...")
    focus_linkedin_chrome(maximize=True)

    scroll_page(config)

    saved_dir = ROOT / "saved_pages"
    saved_dir.mkdir(exist_ok=True)
    before_save = snapshot_saved_pages(saved_dir)

    # Chrome может сохранить под стандартным именем "Лента _ LinkedIn",
    # "Вакансии _ LinkedIn" и т.п. Поэтому не полагаемся на output из диалога.
    output_hint = save_page(target if target in URLS else "page", config)

    print("Жду фактически сохранённый файл...")
    actual_saved = wait_for_saved_page_change(saved_dir, before_save, timeout=18.0)

    if actual_saved is not None:
        actual_saved = rename_saved_page_for_target(
            actual_saved,
            target if target in URLS else "page",
        )
        print(f"Файл текущего запуска: {actual_saved}")

        if config.get("parse_after_save", True):
            time.sleep(1)
            parse_single_file(actual_saved)
    else:
        print("Не удалось определить новый сохранённый файл.")
        print(f"Ожидаемое имя по старой логике было: {output_hint}")

    print("Закрываю текущую вкладку LinkedIn мышью...")
    focus_linkedin_chrome(maximize=True)
    time.sleep(0.6)

    width, height = pyautogui.size()

    # По текущему расположению Chrome LinkedIn открывается второй вкладкой.
    # На скриншоте 1536x958 крестик второй вкладки находится около (527, 21).
    # Используем относительные координаты, чтобы работало и при другом разрешении.
    close_x = int(width * ACTIVE_DEVICE["close_tab_x"])
    close_y = int(height * ACTIVE_DEVICE["close_tab_y"])

    pyautogui.moveTo(close_x, close_y, duration=0.25)
    time.sleep(0.2)
    pyautogui.click()
    time.sleep(0.8)

    print(f"Готово. Последнее сохранение: {actual_saved or output_hint}")


def decode_html_bytes(payload: bytes, declared_charset: str = "") -> str:
    # Chrome LinkedIn MHTML фактически хранит HTML в UTF-8,
    # даже когда charset в MIME-части не указан.
    if not payload:
        return ""

    head = payload[:8192]
    meta_match = re.search(
        br'charset\s*=\s*["\']?([A-Za-z0-9._-]+)',
        head,
        flags=re.I,
    )

    meta_charset = ""
    if meta_match:
        try:
            meta_charset = meta_match.group(1).decode("ascii", errors="ignore")
        except Exception:
            pass

    candidates = []
    for enc in ("utf-8", declared_charset, meta_charset, "windows-1251", "cp1252"):
        enc = (enc or "").strip()
        if enc and enc.lower() not in [x.lower() for x in candidates]:
            candidates.append(enc)

    best_text = ""
    best_score = None

    for enc in candidates:
        try:
            decoded = payload.decode(enc, errors="replace")
        except Exception:
            continue

        replacements = decoded.count("\ufffd")
        mojibake = sum(decoded.count(x) for x in ("Р", "С", "Ð", "Ñ"))
        score = replacements * 1000 + mojibake

        if best_score is None or score < best_score:
            best_score = score
            best_text = decoded

        # Chrome LinkedIn MHTML почти всегда UTF-8. В большом документе могут
        # встретиться единичные повреждённые байты из встроенных ресурсов.
        # Не переключаем из-за них весь документ на windows-1251.
        if enc.lower().replace("_", "-") == "utf-8":
            replacement_ratio = replacements / max(len(decoded), 1)
            if replacement_ratio < 0.001:
                return decoded

    return best_text or payload.decode("utf-8", errors="replace")

def read_html_or_mhtml(path: Path) -> str:
    if path.suffix.lower() in {".mhtml", ".mht"}:
        with path.open("rb") as fh:
            message = BytesParser(policy=policy.default).parse(fh)

        parts = []
        feed_parts = []

        if message.is_multipart():
            for part in message.walk():
                if part.get_content_type() != "text/html":
                    continue

                payload = part.get_payload(decode=True) or b""
                decoded = decode_html_bytes(
                    payload,
                    part.get_content_charset() or "",
                )
                parts.append(decoded)

                content_location = (part.get("Content-Location") or "").strip()
                if re.match(
                    r"https://(?:www\\.)?linkedin\\.com/feed/?(?:[?#].*)?$",
                    content_location,
                    flags=re.IGNORECASE,
                ):
                    feed_parts.append(decoded)
        else:
            payload = message.get_payload(decode=True) or b""
            decoded = decode_html_bytes(
                payload,
                message.get_content_charset() or "",
            )
            parts.append(decoded)

        # Для Feed MHTML явно выбираем основной документ по Content-Location.
        # Это исключает LinkedIn preload/iframe/служебные HTML.
        if feed_parts:
            return max(feed_parts, key=len)

        if not parts:
            return ""

        # Для Jobs/Notifications и прочих страниц сохраняем безопасный fallback.
        return max(parts, key=len)

    return decode_html_bytes(path.read_bytes())

def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def normalize_url(href: str) -> str:
    href = (href or "").strip()
    if href.startswith("/"):
        return "https://www.linkedin.com" + href
    return href


def classify(url: str) -> str:
    value = url.lower()
    if "/jobs/" in value:
        return "job"
    if "/posts/" in value or "/feed/" in value or "/activity-" in value:
        return "post"
    if "/company/" in value:
        return "company"
    if "/in/" in value:
        return "person"
    return "link"


def extract_records(html: str, source_file: str) -> list[dict]:
    """Feed parser: сначала реальные публикации, затем полезные entity-ссылки."""
    if BeautifulSoup is None:
        raise RuntimeError("Не установлен beautifulsoup4.")

    soup = BeautifulSoup(html, "lxml")
    records = []
    seen = set()

    def strip_ui_tail(value: str) -> str:
        value = clean_text(value)
        for suffix in ("… развернуть", "Показать перевод"):
            value = value.replace(suffix, " ")
        return clean_text(value)

    def first_author_link(container):
        # Company post links обычно выглядят /company/<slug>/posts/.
        for a in container.find_all("a", href=True):
            url = normalize_url(a.get("href"))
            if not url:
                continue
            if "/company/" in url and "/posts/" in url:
                title = clean_text(a.get_text(" ", strip=True))
                if title:
                    return title, url, "company"

        # Для публикаций людей берём первый содержательный /in/ link.
        for a in container.find_all("a", href=True):
            url = normalize_url(a.get("href"))
            if not url or "/in/" not in url:
                continue
            title = clean_text(a.get_text(" ", strip=True))
            if title:
                return title, url, "person"

        return "", "", ""

    # ------------------------------------------------------------------
    # 1. Реальные посты нового LinkedIn Feed.
    # У актуального DOM текст публикации находится в expandable-text-box,
    # а карточка поста — ближайший role=listitem.
    # ------------------------------------------------------------------
    post_nodes = soup.select('[data-testid="expandable-text-box"]')

    for body in post_nodes:
        post_text = strip_ui_tail(body.get_text(" ", strip=True))
        if len(post_text) < 20:
            continue

        container = body.find_parent(attrs={"role": "listitem"})
        if container is None:
            container = body.find_parent("article")
        if container is None:
            container = body.parent

        author, author_url, author_type = first_author_link(container)

        # Хэштеги.
        hashtags = []
        for a in body.find_all("a", href=True):
            label = clean_text(a.get_text(" ", strip=True))
            href = a.get("href") or ""
            if label.startswith("#") or "HASH_TAG_FROM_FEED" in href:
                if label and label not in hashtags:
                    hashtags.append(label)

        # Упомянутые LinkedIn-компании/люди внутри текста поста.
        mentions = []
        for a in body.find_all("a", href=True):
            label = clean_text(a.get_text(" ", strip=True))
            url = normalize_url(a.get("href"))
            if not label or not url:
                continue
            if "/company/" in url or "/in/" in url:
                item = {"name": label[:300], "url": url}
                if item not in mentions:
                    mentions.append(item)

        # Внешние/редирект-ссылки, присутствующие в карточке.
        external_links = []
        for a in container.find_all("a", href=True):
            url = normalize_url(a.get("href"))
            if not url:
                continue
            if (
                "linkedin.com/" not in url
                or "lnkd.in/" in url
            ):
                if url not in external_links:
                    external_links.append(url)

        # Время публикации обычно находится рядом с автором в верхней
        # части карточки: "2 дн.", "4 ч.", "2 нед." и т.п.
        header_text = clean_text(container.get_text(" ", strip=True))[:700]
        time_match = re.search(
            r"(?<!\d)(\d+\s*(?:мин\.?|ч\.?|дн\.?|нед\.?|мес\.?|г\.?))(?:\s*[•·])?",
            header_text,
            flags=re.IGNORECASE,
        )
        posted_time = time_match.group(1) if time_match else ""

        # Реакции и комментарии — если LinkedIn успел их загрузить.
        # Считываем отдельный UI-элемент, чтобы цифры из даты/видео/текста
        # не склеивались с метрикой.
        reactions = None
        comments = None

        for metric_node in container.find_all(["a", "button"]):
            metric_text = clean_text(metric_node.get_text(" ", strip=True))
            if not metric_text:
                continue

            if reactions is None and re.search(r"\bреакц", metric_text, re.IGNORECASE):
                m = re.search(r"([\d\s\u00a0.,]+)\s+реакц", metric_text, re.IGNORECASE)
                if m:
                    raw = re.sub(r"[^\d]", "", m.group(1))
                    if raw:
                        reactions = int(raw)

            if comments is None and re.search(r"\bкомментар", metric_text, re.IGNORECASE):
                m = re.search(r"([\d\s\u00a0.,]+)\s+комментар", metric_text, re.IGNORECASE)
                if m:
                    raw = re.sub(r"[^\d]", "", m.group(1))
                    if raw:
                        comments = int(raw)

            if reactions is not None and comments is not None:
                break

        # Сохраняем внутренний идентификатор UGC, если он есть в DOM.
        post_id = ""
        component_blob = " ".join(
            str(x.get("componentkey", ""))
            for x in container.find_all(attrs={"componentkey": True})
        )
        id_match = re.search(r"userGeneratedContentId=(\d+)", component_blob)
        if id_match:
            post_id = id_match.group(1)

        # Если отдельного permalink в DOM нет, author posts/profile URL лучше,
        # чем случайная ссылка из карточки.
        post_url = ""
        for a in container.find_all("a", href=True):
            url = normalize_url(a.get("href"))
            if not url:
                continue
            if "/feed/update/" in url or "/posts/" in url and "activity-" in url:
                post_url = url
                break
        if not post_url:
            post_url = author_url

        key = (
            post_id or post_url.split("?")[0],
            post_text[:500],
        )
        if key in seen:
            continue
        seen.add(key)

        records.append({
            "type": "post",
            "title": (author or post_text[:180])[:500],
            "company": author if author_type == "company" else "",
            "location": "",
            "text": post_text[:30000],
            "url": post_url,
            "source_file": source_file,
            "author": author,
            "author_type": author_type,
            "author_url": author_url,
            "posted_time": posted_time,
            "post_id": post_id,
            "hashtags": hashtags,
            "mentions": mentions,
            "external_links": external_links,
            "reactions": reactions,
            "comments": comments,
        })

    # ------------------------------------------------------------------
    # 2. Полезные company/person entity-ссылки.
    # Они нужны будущему networking/company intelligence, но не выдаём
    # company-post links повторно как отдельные "post" записи.
    # ------------------------------------------------------------------
    for a in soup.find_all("a", href=True):
        url = normalize_url(a.get("href"))
        title = clean_text(a.get_text(" ", strip=True))
        if not url or len(title) < 4:
            continue

        entity_type = None
        canonical_url = url

        if "/company/" in url:
            entity_type = "company"
            # /company/x/posts/ -> canonical company page
            canonical_url = re.sub(r"/posts/?(?:\?.*)?$", "/", url)
        elif "/in/" in url:
            entity_type = "person"
        else:
            continue

        # Убираем очевидные UI-подписи.
        if title.lower() in {
            "отслеживать", "подписаться", "подробнее",
            "показать перевод", "см. все"
        }:
            continue

        key = (canonical_url.split("?")[0], title[:300])
        if key in seen:
            continue
        seen.add(key)

        records.append({
            "type": entity_type,
            "title": title[:500],
            "company": "",
            "location": "",
            "text": title[:2000],
            "url": canonical_url,
            "source_file": source_file,
        })

    return records

def extract_jobs_records(html: str, source_file: str) -> list[dict]:
    """
    Специализированный парсер LinkedIn Jobs.

    Главное правило:
      одна JSON-запись = один уникальный LinkedIn job_id.

    Контейнер, содержащий ссылки сразу на несколько разных вакансий,
    никогда не используется как text/title/company/location одной вакансии.
    Это предотвращает склейку соседних карточек.
    """
    if BeautifulSoup is None:
        raise RuntimeError("Не установлен beautifulsoup4.")

    soup = BeautifulSoup(html, "lxml")
    records = []
    seen_ids = set()

    def get_job_id(url: str) -> str:
        if not url:
            return ""

        match = re.search(r"/jobs/view/(\d+)", url)
        if match:
            return match.group(1)

        try:
            query = parse_qs(urlparse(url).query)
            values = query.get("currentJobId", [])
            if values and str(values[0]).isdigit():
                return str(values[0])
        except Exception:
            pass

        match = re.search(r"[?&]currentJobId=(\d+)", url)
        return match.group(1) if match else ""

    def job_ids_in_node(node) -> set[str]:
        ids = set()
        try:
            links = node.find_all("a", href=True)
        except Exception:
            return ids

        for a in links:
            jid = get_job_id(normalize_url(a.get("href")))
            if jid:
                ids.add(jid)
        return ids

    def first_text(node, selectors):
        for selector in selectors:
            try:
                found = node.select_one(selector)
            except Exception:
                found = None
            if found:
                value = clean_text(found.get_text(" ", strip=True))
                if value:
                    return value
        return ""

    def text_from_job_link(a) -> str:
        value = clean_text(a.get_text(" ", strip=True))
        if not value:
            return ""

        value = re.sub(
            r"\s*\((?:подтвержденная вакансия|verified job)\)\s*",
            " ",
            value,
            flags=re.I,
        )
        return clean_text(value)[:500]

    def choose_single_job_container(anchor, job_id: str):
        """
        Идём от самой ссылки вверх и выбираем НАИБОЛЬШИЙ разумный контейнер,
        который всё ещё содержит только текущий job_id.

        Как только родитель начинает содержать несколько вакансий,
        подъём прекращается.
        """
        chosen = anchor
        node = anchor

        for _ in range(9):
            parent = getattr(node, "parent", None)
            if parent is None:
                break

            ids_here = job_ids_in_node(parent)

            # Вышли на общий список/панель нескольких вакансий.
            if len(ids_here) > 1:
                break

            # Если контейнер содержит только текущую вакансию, он безопасен.
            if not ids_here or ids_here == {job_id}:
                candidate_text = clean_text(parent.get_text(" ", strip=True))

                # Не берём огромные layout-контейнеры даже при одном ID.
                if 8 <= len(candidate_text) <= 6000:
                    chosen = parent

            node = parent

        return chosen

    def parse_link_payload(job_link, title, company, location):
        """
        В MHTML LinkedIn текст ссылки нередко содержит:
        title + verified marker + company • location ...
        Используем это только как fallback.
        """
        if job_link is None:
            return title, company, location

        link_text = clean_text(job_link.get_text(" ", strip=True))
        if not link_text:
            return title, company, location

        marker_match = re.search(
            r"\s*\((?:подтвержденная вакансия|verified job)\)\s*",
            link_text,
            flags=re.I,
        )

        if marker_match:
            clean_title = clean_text(link_text[:marker_match.start()])
            rest = clean_text(link_text[marker_match.end():])

            if clean_title:
                title = clean_title[:500]

            if title and rest.startswith(title):
                rest = clean_text(rest[len(title):])

            parts = [clean_text(x) for x in rest.split("•") if clean_text(x)]

            if parts:
                company = parts[0][:500]
                if len(parts) >= 2:
                    location = parts[1][:500]

        elif title and link_text.startswith(title):
            rest = clean_text(link_text[len(title):])

            # Иногда title дублируется внутри anchor.
            if rest.startswith(title):
                rest = clean_text(rest[len(title):])

            parts = [clean_text(x) for x in rest.split("•") if clean_text(x)]

            if parts:
                if not company:
                    company = parts[0][:500]
                if len(parts) >= 2 and not location:
                    location = parts[1][:500]

        return title, company, location

    def add_job(node, url="", force_job_id=""):
        job_id = force_job_id or get_job_id(url)

        if not job_id:
            # Разрешаем автоматический поиск URL только если контейнер
            # относится ровно к одной вакансии.
            ids_here = job_ids_in_node(node)
            if len(ids_here) != 1:
                return

            job_id = next(iter(ids_here))

            for a in node.find_all("a", href=True):
                candidate = normalize_url(a.get("href"))
                if get_job_id(candidate) == job_id:
                    url = candidate
                    break

        if not job_id or job_id in seen_ids:
            return

        # Критическая защита от склейки карточек.
        ids_here = job_ids_in_node(node)
        if len(ids_here) > 1:
            return
        if ids_here and ids_here != {job_id}:
            return

        full_text = clean_text(node.get_text(" ", strip=True))
        if len(full_text) < 8:
            return

        job_link = None
        for a in node.find_all("a", href=True):
            candidate = normalize_url(a.get("href"))
            if get_job_id(candidate) == job_id:
                job_link = a
                if not url:
                    url = candidate
                break

        title = first_text(node, [
            ".job-card-list__title",
            ".job-card-container__link",
            ".jobs-unified-top-card__job-title",
            ".job-details-jobs-unified-top-card__job-title",
            "[class*='job-title']",
            "h1", "h2", "h3",
        ])

        company = first_text(node, [
            ".job-card-container__primary-description",
            ".artdeco-entity-lockup__subtitle",
            ".jobs-unified-top-card__company-name",
            ".job-details-jobs-unified-top-card__company-name",
            "[class*='company-name']",
            "a[href*='/company/']",
        ])

        location = first_text(node, [
            ".job-card-container__metadata-item",
            ".jobs-unified-top-card__bullet",
            "[class*='job-card-container__metadata']",
            "[class*='primary-description-container']",
        ])

        if not title and job_link is not None:
            title = text_from_job_link(job_link)

        title, company, location = parse_link_payload(
            job_link, title, company, location
        )

        # Fallback для карточек, где LinkedIn кладёт в anchor весь текст:
        # "Title Title Company • Location ...".
        # В таком случае старый селектор может ошибочно принять весь anchor
        # за title, оставив company/location пустыми.
        if (not company or not location) and job_link is not None:
            link_text = clean_text(job_link.get_text(" ", strip=True))
            if "•" in link_text:
                left, right = [clean_text(x) for x in link_text.split("•", 1)]
                words = left.split()

                # Ищем повторяющийся title в начале:
                # [Marketing Manager] [Marketing Manager] [SKYWORTH]
                for k in range(1, len(words) // 2 + 1):
                    if words[:k] == words[k:2 * k]:
                        recovered_title = clean_text(" ".join(words[:k]))
                        recovered_company = clean_text(" ".join(words[2 * k:]))

                        if recovered_title and recovered_company:
                            title = recovered_title[:500]
                            if not company:
                                company = recovered_company[:500]
                            if not location:
                                location = right[:500]
                            break

        if not title:
            # Последний fallback, но только из уже безопасного single-job node.
            title = full_text[:180]

        # Дополнительная sanity-check:
        # если каким-то образом text всё же содержит ссылки на другой job_id,
        # запись не публикуем.
        other_ids = job_ids_in_node(node) - {job_id}
        if other_ids:
            return

        seen_ids.add(job_id)

        # Стабильная canonical URL полезнее currentJobId-параметров.
        canonical_url = f"https://www.linkedin.com/jobs/view/{job_id}/"

        records.append({
            "type": "job",
            "job_id": job_id,
            "title": title[:500],
            "company": company[:500],
            "location": location[:500],
            "text": full_text[:15000],
            "url": canonical_url,
            "source_file": source_file,
        })

    # 1) Сначала самые надёжные job-card контейнеры.
    safe_selectors = [
        ".jobs-search-results__list-item",
        ".job-card-container",
        ".job-card-list",
        "li[data-occludable-job-id]",
        "[data-job-id]",
    ]

    for selector in safe_selectors:
        try:
            nodes = soup.select(selector)
        except Exception:
            continue

        for node in nodes:
            ids_here = job_ids_in_node(node)

            # Только один уникальный job_id.
            if len(ids_here) == 1:
                add_job(node)

    # 2) Fallback по каждой ссылке на вакансию.
    # Для неё ищем ближайший безопасный контейнер, в котором нет соседних jobs.
    for a in soup.find_all("a", href=True):
        url = normalize_url(a.get("href"))
        job_id = get_job_id(url)

        if not job_id or job_id in seen_ids:
            continue

        chosen = choose_single_job_container(a, job_id)
        add_job(chosen, url=url, force_job_id=job_id)

    return records

def rename_linkedin_saved_pages(folder: Path) -> None:
    """Переименовывает стандартные имена Chrome в наши понятные имена."""
    folder.mkdir(exist_ok=True)

    mappings = {
        "Лента _ LinkedIn": "feed",
        "Вакансии _ LinkedIn": "jobs",
        "Уведомления _ LinkedIn": "notifications",
        "Feed _ LinkedIn": "feed",
        "Jobs _ LinkedIn": "jobs",
        "Notifications _ LinkedIn": "notifications",
    }

    for path in list(folder.iterdir()):
        if not path.is_file():
            continue
        if path.suffix.lower() not in {".mhtml", ".mht", ".html", ".htm"}:
            continue

        target_type = mappings.get(path.stem)
        if not target_type:
            continue

        # Используем время изменения самого сохранённого файла.
        stamp = datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d_%H-%M-%S")
        new_path = folder / f"linkedin_{target_type}_{stamp}{path.suffix.lower()}"

        # Если такое имя уже есть, добавляем порядковый номер.
        counter = 2
        while new_path.exists():
            new_path = folder / f"linkedin_{target_type}_{stamp}_{counter}{path.suffix.lower()}"
            counter += 1

        path.rename(new_path)
        print(f"Переименован файл: {path.name} -> {new_path.name}")




def clean_notification_text(value: str) -> str:
    """Убирает характерные повторения текста внутри карточек LinkedIn Notifications."""
    value = clean_text(value)
    if not value:
        return value

    # 1) Если вся строка состоит из 2-4 одинаковых подряд идущих частей.
    for parts in (4, 3, 2):
        if len(value) % parts == 0:
            piece_len = len(value) // parts
            pieces = [value[i * piece_len:(i + 1) * piece_len].strip() for i in range(parts)]
            if pieces and len(pieces[0]) >= 15 and all(x == pieces[0] for x in pieces[1:]):
                value = pieces[0]
                break

    # 2) Удаляем подряд идущие одинаковые фразы/фрагменты после разделения.
    chunks = re.split(r"(?<=[.!?…])\s+|\s{2,}", value)
    compact = []
    previous = None
    for chunk in chunks:
        chunk = chunk.strip()
        if not chunk:
            continue
        normalized = re.sub(r"\s+", " ", chunk).casefold()
        if normalized == previous:
            continue
        compact.append(chunk)
        previous = normalized

    if compact:
        value = " ".join(compact)

    # 3) Частый MHTML-случай: одинаковая длинная последовательность повторилась подряд.
    # Ищем максимальный префикс, который сразу же повторяется.
    max_len = min(len(value) // 2, 1200)
    for size in range(max_len, 30, -1):
        first = value[:size].strip()
        second = value[size:size * 2].strip()
        if first == second:
            value = first + value[size * 2:]
            break

    return clean_text(value)


def extract_notification_records(html: str, source_file: str) -> list[dict]:
    """Извлекает реальные headline-карточки LinkedIn Notifications."""
    if BeautifulSoup is None:
        raise RuntimeError("Не установлен beautifulsoup4.")

    soup = BeautifulSoup(html, "lxml")
    records = []
    seen = set()

    def useful_url(url: str) -> bool:
        low = (url or "").lower()
        return any(part in low for part in (
            "/feed/",
            "/in/",
            "/company/",
            "/posts/",
            "/activity-",
            "/jobs/",
            "/suggested-for-you/",
        ))

    # Реальная текущая разметка сохранённого Notifications:
    # componentkey="notification-card-headline_<UUID>"
    headlines = soup.select("[componentkey^='notification-card-headline_']")

    for headline in headlines:
        component_key = headline.get("componentkey", "")
        text_value = clean_notification_text(
            headline.get_text(" ", strip=True)
        )

        if len(text_value) < 8:
            continue

        key = component_key or text_value[:500]
        if key in seen:
            continue
        seen.add(key)

        container = headline
        best_links = []

        for _ in range(6):
            parent = getattr(container, "parent", None)
            if parent is None:
                break
            container = parent

            links = [
                normalize_url(a.get("href"))
                for a in container.find_all("a", href=True)
            ]
            links = [x for x in links if x]

            if links:
                best_links = links

            ctext = clean_text(container.get_text(" ", strip=True))
            if links and len(links) <= 5 and len(ctext) <= max(5000, len(text_value) * 3):
                break

        url = next((x for x in best_links if useful_url(x)), "")
        if not url and best_links:
            url = best_links[0]

        records.append({
            "type": "notification",
            "title": text_value[:220],
            "company": "",
            "location": "",
            "text": text_value[:15000],
            "url": url,
            "source_file": source_file,
        })

    return records


def parse_folder(folder: Path, config_path: Path) -> None:
    # Сначала приводим стандартные имена Chrome к нашим именам.
    rename_linkedin_saved_pages(folder)

    files = sorted(
        p for p in folder.rglob("*")
        if p.suffix.lower() in {".mhtml", ".mht", ".html", ".htm"}
    )
    if not files:
        print(f"В папке {folder} пока нет MHTML/HTML.")
        return

    all_items = []
    for path in files:
        print(f"Парсинг: {path.name}")
        try:
            html = read_html_or_mhtml(path)

            low_name = path.name.lower()
            if "notification" in low_name or "уведом" in low_name:
                items = extract_notification_records(html, path.name)
                print(f"  Notifications: найдено карточек {len(items)}")
            elif "linkedin_jobs_" in low_name or "ваканс" in low_name or "jobs _ linkedin" in low_name:
                items = extract_jobs_records(html, path.name)
                print(f"  Jobs: найдено вакансий {len(items)}")
            else:
                items = extract_records(html, path.name)

            all_items.extend(items)
        except Exception as exc:
            all_items.append({
                "type": "error",
                "title": str(exc),
                "company": "",
                "location": "",
                "text": "",
                "url": "",
                "source_file": path.name,
            })

    deduped = []
    seen = set()
    for item in all_items:
        key = (
            item.get("url", "").split("?")[0],
            item.get("title", "")[:300],
            item.get("text", "")[:300],
        )
        if key in seen:
            continue
        seen.add(key)
        item["collected_at"] = datetime.now().isoformat(timespec="seconds")
        deduped.append(item)

    deduped.sort(key=lambda x: (x.get("type", ""), x.get("title", "")))

    reports = ROOT / "reports"
    reports.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    base = reports / f"linkedin_report_{stamp}"

    base.with_suffix(".json").write_text(
        json.dumps(deduped, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    columns = [
        "type", "job_id", "title", "company", "location",
        "url", "source_file", "collected_at", "text"
    ]
    with base.with_suffix(".csv").open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(deduped)

    if pd is not None:
        df = pd.DataFrame(deduped)
        with pd.ExcelWriter(base.with_suffix(".xlsx"), engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="All", index=False)
            if not df.empty:
                for kind in ("job", "post", "notification", "company", "person", "text", "error"):
                    part = df[df["type"] == kind]
                    if not part.empty:
                        part.to_excel(writer, sheet_name=kind, index=False)
        print(f"Excel: {base.with_suffix('.xlsx')}")
    else:
        print("pandas не установлен — созданы JSON и CSV, Excel пропущен.")

    print(f"JSON: {base.with_suffix('.json')}")
    print(f"CSV: {base.with_suffix('.csv')}")
    print(f"Записей: {len(deduped)}")



def parse_period(folder: Path, config_path: Path) -> None:
    """Парсит только сохранённые страницы, чья дата попадает в заданный период."""
    rename_linkedin_saved_pages(folder)

    start_raw = input("Начало периода (YYYY-MM-DD): ").strip()
    end_raw = input("Конец периода   (YYYY-MM-DD): ").strip()

    try:
        start_date = datetime.strptime(start_raw, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_raw, "%Y-%m-%d").date()
    except ValueError:
        print("Неверный формат даты. Используйте YYYY-MM-DD.")
        return

    if end_date < start_date:
        print("Конечная дата не может быть раньше начальной.")
        return

    files = sorted(
        p for p in folder.rglob("*")
        if p.suffix.lower() in {".mhtml", ".mht", ".html", ".htm"}
    )

    selected = []
    for path in files:
        try:
            file_date = datetime.fromtimestamp(path.stat().st_mtime).date()
        except OSError:
            continue

        if start_date <= file_date <= end_date:
            selected.append(path)

    if not selected:
        print(f"За период {start_date} ... {end_date} файлов не найдено.")
        return

    print(f"Файлов за период: {len(selected)}")

    all_items = []
    for path in selected:
        print(f"Парсинг: {path.name}")
        try:
            html = read_html_or_mhtml(path)

            low_name = path.name.lower()
            if "notification" in low_name or "уведом" in low_name:
                items = extract_notification_records(html, path.name)
                print(f"  Notifications: найдено карточек {len(items)}")
            elif "linkedin_jobs_" in low_name or "ваканс" in low_name or "jobs _ linkedin" in low_name:
                items = extract_jobs_records(html, path.name)
                print(f"  Jobs: найдено вакансий {len(items)}")
            else:
                items = extract_records(html, path.name)

            all_items.extend(items)
        except Exception as exc:
            all_items.append({
                "type": "error",
                "title": str(exc),
                "company": "",
                "location": "",
                "text": "",
                "url": "",
                "source_file": path.name,
            })

    deduped = []
    seen = set()
    for item in all_items:
        key = (
            item.get("url", "").split("?")[0],
            item.get("title", "")[:300],
            item.get("text", "")[:300],
        )
        if key in seen:
            continue
        seen.add(key)
        item["collected_at"] = datetime.now().isoformat(timespec="seconds")
        deduped.append(item)

    deduped.sort(key=lambda x: (x.get("type", ""), x.get("title", "")))

    reports = ROOT / "reports"
    reports.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    base = reports / f"linkedin_period_{start_date}_to_{end_date}_{stamp}"

    base.with_suffix(".json").write_text(
        json.dumps(deduped, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    columns = [
        "type", "job_id", "title", "company", "location",
        "url", "source_file", "collected_at", "text"
    ]

    with base.with_suffix(".csv").open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(deduped)

    if pd is not None:
        df = pd.DataFrame(deduped)
        with pd.ExcelWriter(base.with_suffix(".xlsx"), engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="All", index=False)
            if not df.empty:
                for kind in ("job", "post", "notification", "company", "person", "text", "error"):
                    part = df[df["type"] == kind]
                    if not part.empty:
                        part.to_excel(writer, sheet_name=kind, index=False)
        print(f"Excel: {base.with_suffix('.xlsx')}")

    print(f"JSON: {base.with_suffix('.json')}")
    print(f"CSV: {base.with_suffix('.csv')}")
    print(f"Записей: {len(deduped)}")




def newest_target_csv(reports_dir: Path, target: str, after_timestamp: float = 0.0) -> Path | None:
    """Возвращает самый свежий CSV-отчёт нужного типа, созданный после указанного времени."""
    candidates = []
    pattern = f"linkedin_{target}_report_*.csv"

    for path in reports_dir.glob(pattern):
        try:
            mtime = path.stat().st_mtime
        except OSError:
            continue

        if mtime >= after_timestamp:
            candidates.append((mtime, path))

    if not candidates:
        return None

    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def create_combined_sequence_report(report_files: dict[str, Path]) -> Path | None:
    """Создаёт один Excel с листами Summary, Feed, Jobs, Notifications."""
    if pd is None:
        print("Не удалось создать общий Excel: не установлен pandas.")
        return None

    reports_dir = ROOT / "reports"
    reports_dir.mkdir(exist_ok=True)

    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output = reports_dir / f"linkedin_combined_report_{stamp}.xlsx"

    sheet_names = {
        "feed": "Feed",
        "jobs": "Jobs",
        "notifications": "Notifications",
    }

    loaded = {}
    summary_rows = []

    for target in ("feed", "jobs", "notifications"):
        path = report_files.get(target)

        if path is None or not path.exists():
            loaded[target] = pd.DataFrame()
            summary_rows.append({
                "section": sheet_names[target],
                "records": 0,
                "source_report": "",
                "status": "не найден",
            })
            continue

        try:
            df = pd.read_csv(path, encoding="utf-8-sig")
        except Exception:
            try:
                df = pd.read_csv(path, encoding="utf-8")
            except Exception as exc:
                print(f"Не удалось прочитать {path.name}: {exc}")
                df = pd.DataFrame()

        loaded[target] = df
        summary_rows.append({
            "section": sheet_names[target],
            "records": len(df),
            "source_report": path.name,
            "status": "OK",
        })

    summary_df = pd.DataFrame(summary_rows)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        for target in ("feed", "jobs", "notifications"):
            df = loaded[target]
            df.to_excel(
                writer,
                sheet_name=sheet_names[target],
                index=False,
            )

        # Немного улучшаем читаемость итогового файла.
        workbook = writer.book
        for ws in workbook.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for column_cells in ws.columns:
                max_len = 0
                column_letter = column_cells[0].column_letter

                for cell in column_cells[:100]:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(value), 60))

                ws.column_dimensions[column_letter].width = max(10, min(max_len + 2, 60))

    print("\nОбщий отчёт последовательного запуска:")
    print(f"  {output}")
    print(
        "  Feed: "
        f"{len(loaded['feed'])}, "
        "Jobs: "
        f"{len(loaded['jobs'])}, "
        "Notifications: "
        f"{len(loaded['notifications'])}"
    )

    return output



def _get_or_create_worksheet(spreadsheet, title: str, rows: int = 1000, cols: int = 20):
    """Возвращает существующий лист или создаёт новый."""
    try:
        return spreadsheet.worksheet(title)
    except Exception:
        return spreadsheet.add_worksheet(title=title, rows=rows, cols=cols)


def _append_dataframe_to_worksheet(ws, df, extra_columns: dict | None = None) -> int:
    """Добавляет DataFrame в конец листа, сохраняя историю."""
    if df is None or df.empty:
        return 0

    data = df.copy()

    if extra_columns:
        for key, value in extra_columns.items():
            data[key] = value

    data = data.fillna("")

    headers = [str(x) for x in data.columns]
    rows = [
        [
            x.isoformat() if hasattr(x, "isoformat") else str(x)
            for x in row
        ]
        for row in data.itertuples(index=False, name=None)
    ]

    existing = ws.get_all_values()

    if not existing:
        ws.append_row(headers, value_input_option="RAW")
    else:
        current_headers = existing[0]
        if current_headers != headers:
            # Если структура изменилась, не портим существующий лист:
            # добавляем недостающие колонки в заголовок справа.
            merged = list(current_headers)
            for h in headers:
                if h not in merged:
                    merged.append(h)

            if merged != current_headers:
                ws.update("A1", [merged])

            # Приводим строки к текущему порядку колонок.
            index_map = {name: i for i, name in enumerate(headers)}
            normalized_rows = []
            for row in rows:
                normalized_rows.append([
                    row[index_map[h]] if h in index_map else ""
                    for h in merged
                ])
            rows = normalized_rows

    if rows:
        ws.append_rows(rows, value_input_option="RAW")

    return len(rows)


def json_safe(value):
    """Преобразует значения Excel в JSON-совместимый вид."""
    if value is None:
        return None
    if isinstance(value, (datetime, date, dt_time)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)



def worksheet_to_json(ws_formula, ws_value):
    """
    Сохраняет:
      - records: строки по названиям колонок
      - raw_values: полную матрицу значений
      - formulas: формулы по координатам
      - cached_formula_values: сохранённые Excel-значения формул
      - merged_ranges: объединённые диапазоны
    """
    max_row = ws_formula.max_row
    max_col = ws_formula.max_column

    raw_values = []
    formulas = {}
    cached_formula_values = {}

    for row in range(1, max_row + 1):
        out_row = []
        for col in range(1, max_col + 1):
            cell_formula = ws_formula.cell(row=row, column=col)
            cell_value = ws_value.cell(row=row, column=col)

            value = cell_formula.value
            if cell_formula.data_type == "f":
                coord = cell_formula.coordinate
                formulas[coord] = json_safe(value)
                cached_formula_values[coord] = json_safe(cell_value.value)

                # В raw_values кладём саму формулу, чтобы её не потерять.
                out_row.append(json_safe(value))
            else:
                out_row.append(json_safe(value))

        raw_values.append(out_row)

    headers = raw_values[0] if raw_values else []

    columns = []
    used = set()

    for index, header in enumerate(headers, start=1):
        if header in (None, ""):
            name = f"column_{index}"
        else:
            name = str(header)

        base = name
        counter = 2
        while name in used:
            name = f"{base}_{counter}"
            counter += 1

        used.add(name)
        columns.append(name)

    records = []
    for row in raw_values[1:]:
        record = {}
        for index, column_name in enumerate(columns):
            record[column_name] = row[index] if index < len(row) else None
        records.append(record)

    merged_ranges = [str(x) for x in ws_formula.merged_cells.ranges]

    return {
        "title": ws_formula.title,
        "max_row": max_row,
        "max_column": max_col,
        "columns": columns,
        "records": records,
        "raw_values": raw_values,
        "formulas": formulas,
        "cached_formula_values": cached_formula_values,
        "merged_ranges": merged_ranges,
    }


def convert_xlsx_to_json(xlsx_path: Path, output_path: Path) -> Path:
    wb_formula = load_workbook(xlsx_path, data_only=False, read_only=False)
    wb_value = load_workbook(xlsx_path, data_only=True, read_only=False)

    result = {
        "schema_version": 1,
        "source": {
            "filename": xlsx_path.name,
            "full_path": str(xlsx_path),
            "modified_at": datetime.fromtimestamp(
                xlsx_path.stat().st_mtime
            ).isoformat(timespec="seconds"),
            "converted_at": datetime.now().isoformat(timespec="seconds"),
        },
        "sheet_order": list(wb_formula.sheetnames),
        "sheets": {},
    }

    for sheet_name in wb_formula.sheetnames:
        ws_formula = wb_formula[sheet_name]
        ws_value = wb_value[sheet_name]
        result["sheets"][sheet_name] = worksheet_to_json(ws_formula, ws_value)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(result, ensure_ascii=False, indent=2, allow_nan=False),
        encoding="utf-8",
    )
    return output_path


def load_github_token() -> str:
    token = os.environ.get("GITHUB_TOKEN", "").strip()
    if token:
        return token

    for path in TOKEN_FILES:
        if path.exists():
            value = path.read_text(encoding="utf-8").strip()
            if value:
                return value

    raise RuntimeError(
        "GitHub token не найден.\n"
        "Варианты:\n"
        "  1) задать переменную окружения GITHUB_TOKEN\n"
        "  2) положить token в файл github_token.txt рядом со скриптом\n"
        "  3) положить token в reports/github_token.txt"
    )


def github_api_request(url: str, token: str, method: str = "GET", data: dict | None = None):
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "linkedin-monitor-uploader",
    }

    payload = None
    if data is not None:
        payload = json.dumps(data).encode("utf-8")
        headers["Content-Type"] = "application/json"

    req = Request(url, data=payload, headers=headers, method=method)
    with urlopen(req, timeout=60) as resp:
        return json.loads(resp.read().decode("utf-8"))


def get_existing_file_sha(token: str, remote_path: str = GITHUB_PATH) -> str | None:
    url = (
        f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/"
        f"{remote_path}?ref={GITHUB_BRANCH}"
    )
    try:
        data = github_api_request(url, token, method="GET")
        return data.get("sha")
    except HTTPError as exc:
        if exc.code == 404:
            return None
        raise


def upload_file_to_github(local_path: Path, remote_path: str = GITHUB_PATH) -> str:
    token = load_github_token()

    raw = local_path.read_bytes()
    content_b64 = base64.b64encode(raw).decode("ascii")

    sha = get_existing_file_sha(token, remote_path)

    api_url = (
        f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/"
        f"{remote_path}"
    )

    message = (
        f"Update {remote_path} from {local_path.name} "
        f"at {datetime.now().isoformat(timespec='seconds')}"
    )

    body = {
        "message": message,
        "content": content_b64,
        "branch": GITHUB_BRANCH,
    }
    if sha:
        body["sha"] = sha

    response = github_api_request(api_url, token, method="PUT", data=body)

    content = response.get("content", {}) or {}
    download_url = content.get("download_url") or (
        f"https://raw.githubusercontent.com/"
        f"{GITHUB_OWNER}/{GITHUB_REPO}/{GITHUB_BRANCH}/{remote_path}"
    )
    return download_url



def publish_combined_to_github(combined_xlsx: Path) -> str:
    """
    Текущий общий Excel -> reports/latest.json -> GitHub/latest.json.
    Использует именно Excel, созданный текущим последовательным запуском.
    """
    if load_workbook is None:
        raise RuntimeError(
            "Не установлен openpyxl. Установите один раз: pip install openpyxl"
        )

    print("\n" + "=" * 62)
    print(" Excel -> latest.json -> GitHub")
    print("=" * 62)
    print(f"Источник: {combined_xlsx.name}")

    result_json = convert_xlsx_to_json(combined_xlsx, LATEST_JSON)
    print(f"JSON создан: {result_json}")
    print(f"Размер JSON: {result_json.stat().st_size:,} байт")

    raw_url = upload_file_to_github(result_json)

    # Ежедневная история: один JSON на дату.
    # Повторный запуск в тот же день обновляет файл этого дня.
    history_path = f"history/{datetime.now().date().isoformat()}.json"
    history_url = upload_file_to_github(result_json, history_path)

    print("GitHub обновлён.")
    print(f"Репозиторий: {GITHUB_OWNER}/{GITHUB_REPO}")
    print(f"Файл: {GITHUB_PATH}")
    print(f"RAW URL: {raw_url}")
    print(f"История: {history_path}")
    print(f"History RAW URL: {history_url}")
    return raw_url

def automate_sequence() -> None:
    """Последовательно запускает Feed -> Jobs -> Notifications и создаёт общий Excel."""
    config = load_config()
    pause_seconds = float(config.get("sequence_pause_seconds", 5))

    targets = [
        ("feed", "Feed"),
        ("jobs", "Jobs"),
        ("notifications", "Notifications"),
    ]

    reports_dir = ROOT / "reports"
    reports_dir.mkdir(exist_ok=True)

    report_files = {}

    print("\n" + "=" * 56)
    print(" Feed -> Jobs -> Notifications + общий Excel")
    print("=" * 56)
    print(f"Пауза между этапами: {pause_seconds:g} сек.")

    for index, (target, label) in enumerate(targets, start=1):
        print(f"\n[{index}/3] Запуск: {label}")

        # Запоминаем момент до запуска, чтобы не взять старый отчёт.
        started_at = time.time()
        automate(target)

        current_report = newest_target_csv(
            reports_dir,
            target,
            after_timestamp=started_at - 1.0,
        )

        if current_report is not None:
            report_files[target] = current_report
            print(f"Отчёт этапа {label}: {current_report.name}")

            # Каждый этап уже создаёт JSON рядом с CSV.
            # Публикуем его отдельным компактным файлом для анализаторов.
            current_json = current_report.with_suffix(".json")
            if current_json.exists():
                remote_latest = f"{target}_latest.json"
                try:
                    section_url = upload_file_to_github(current_json, remote_latest)
                    print(f"GitHub {label}: {remote_latest}")
                    print(f"RAW URL: {section_url}")

                    # Дневная история по каждому разделу.
                    history_section_path = (
                        f"history/{target}/{datetime.now().date().isoformat()}.json"
                    )
                    history_section_url = upload_file_to_github(
                        current_json,
                        history_section_path,
                    )
                    print(f"История {label}: {history_section_path}")
                    print(f"History RAW URL: {history_section_url}")
                except Exception as exc:
                    print(
                        f"ВНИМАНИЕ: не удалось обновить {remote_latest} "
                        f"или его историю: {exc}"
                    )
            else:
                print(
                    f"ВНИМАНИЕ: JSON этапа {label} не найден: "
                    f"{current_json.name}"
                )
        else:
            print(f"Внимание: новый CSV-отчёт этапа {label} не найден.")

        if index < len(targets):
            print(f"\nПауза {pause_seconds:g} сек. перед следующим этапом...")
            time.sleep(pause_seconds)

    combined = create_combined_sequence_report(report_files)

    if combined is not None:
        print(f"\nИтоговый общий файл: {combined.name}")

        try:
            publish_combined_to_github(combined)
        except Exception as exc:
            print("\nВНИМАНИЕ: общий Excel создан, но публикация в GitHub не удалась.")
            print(f"Причина: {exc}")
            print("Excel и локальные отчёты сохранены; их данные не потеряны.")
    else:
        print("\nGitHub-публикация пропущена: общий Excel не был создан.")

    print("\nПоследовательный запуск 1-3 завершён.")


def show_menu() -> None:
    while True:
        print("\n" + "=" * 52)
        print(" LinkedIn Desktop Automation")
        print("=" * 52)
        print("  1 - Feed")
        print("  2 - Jobs")
        print("  3 - Notifications")
        print("  4 - Feed + Jobs + Notifications -> Excel -> JSON -> GitHub")
        print("  5 - Parse за период")
        print("  6 - Parse весь архив")
        print("  7 - Exit")
        print("=" * 52)

        choice = input("Выберите действие [1-7]: ").strip()

        if choice == "1":
            automate("feed")
            input("\nНажмите Enter для возврата в меню...")
        elif choice == "2":
            automate("jobs")
            input("\nНажмите Enter для возврата в меню...")
        elif choice == "3":
            automate("notifications")
            input("\nНажмите Enter для возврата в меню...")
        elif choice == "4":
            automate_sequence()
            input("\nНажмите Enter для возврата в меню...")
        elif choice == "5":
            parse_period(ROOT / "saved_pages", ROOT / "config.json")
            input("\nНажмите Enter для возврата в меню...")
        elif choice == "6":
            parse_folder(ROOT / "saved_pages", ROOT / "config.json")
            input("\nНажмите Enter для возврата в меню...")
        elif choice == "7":
            print("Выход.")
            break
        else:
            print("Неверный выбор. Введите число от 1 до 7.")



# ----------------------------------------------------------------------
# Автоматический режим для компьютера 2
# ----------------------------------------------------------------------
START_TIME_FILE = ROOT / "start.txt"
FINISH_TIME_FILE = ROOT / "finish.txt"
TIME_FILE_POLL_SECONDS = 10


def _read_hhmm_file(path: Path) -> tuple[int, int]:
    """
    Читает время из текстового файла рядом со скриптом.

    Допустимые варианты:
        22 34
        22:34
        22.34
        08 03
        08:03
    """
    if not path.exists():
        raise FileNotFoundError(
            f"Не найден файл времени: {path.name}\n"
            f"Создайте его рядом со скриптом, например: 22 34"
        )

    raw = path.read_text(encoding="utf-8-sig").strip()
    nums = re.findall(r"\d+", raw)

    if len(nums) < 2:
        raise ValueError(
            f"Неверный формат {path.name}: {raw!r}. "
            "Нужно, например: 22 34 или 22:34"
        )

    hour = int(nums[0])
    minute = int(nums[1])

    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        raise ValueError(
            f"Неверное время в {path.name}: {hour:02d}:{minute:02d}"
        )

    return hour, minute


def _today_at(hour: int, minute: int, now: datetime | None = None) -> datetime:
    now = now or datetime.now()
    return now.replace(hour=hour, minute=minute, second=0, microsecond=0)


def _resolve_cycle_times(now: datetime) -> tuple[datetime, datetime]:
    sh, sm = _read_hhmm_file(START_TIME_FILE)
    fh, fm = _read_hhmm_file(FINISH_TIME_FILE)

    start_today = _today_at(sh, sm, now)
    finish_today = _today_at(fh, fm, now)

    # Если finish по часам раньше/equal start — считаем, что finish уже на следующем дне.
    if (fh, fm) <= (sh, sm):
        finish_for_start_today = finish_today + timedelta(days=1)
    else:
        finish_for_start_today = finish_today

    if now < start_today:
        return start_today, finish_for_start_today

    # ПК запустился уже после START, но ещё до FINISH — запускаем цикл сразу.
    if now < finish_for_start_today:
        return now, finish_for_start_today

    # Сегодняшний цикл полностью прошёл — ждём завтра.
    return start_today + timedelta(days=1), finish_for_start_today + timedelta(days=1)


def _wait_until(target: datetime, label: str) -> None:
    last_minute = None

    while True:
        now = datetime.now()
        remaining = (target - now).total_seconds()

        if remaining <= 0:
            print(f"\n[{datetime.now():%Y-%m-%d %H:%M:%S}] {label}")
            return

        minute_key = now.strftime("%Y-%m-%d %H:%M")
        if minute_key != last_minute:
            last_minute = minute_key
            print(
                f"[{now:%H:%M:%S}] Ожидание {label}: "
                f"{target:%Y-%m-%d %H:%M}"
            )

        time.sleep(min(TIME_FILE_POLL_SECONDS, max(1.0, remaining)))


def _shutdown_windows_now() -> None:
    print(f"\n[{datetime.now():%Y-%m-%d %H:%M:%S}] FINISH: выключаем компьютер...")
    subprocess.run(
        ["shutdown", "/s", "/f", "/t", "0"],
        check=False,
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
    )


def unattended_pc2_timed() -> None:
    """
    - без вопросов выбирает профиль компьютера 2;
    - ждёт start.txt;
    - выполняет полный пункт 4;
    - ждёт finish.txt;
    - выключает Windows.
    """
    global ACTIVE_DEVICE
    ACTIVE_DEVICE = DEVICE_PRESETS["2"]

    print("=" * 66)
    print(" LinkedIn AUTO MODE (-a)")
    print("=" * 66)
    print(f"GUI-профиль: {ACTIVE_DEVICE['name']}")
    print(f"start.txt : {START_TIME_FILE}")
    print(f"finish.txt: {FINISH_TIME_FILE}")

    try:
        now = datetime.now()
        run_at, finish_at = _resolve_cycle_times(now)

        sh, sm = _read_hhmm_file(START_TIME_FILE)
        fh, fm = _read_hhmm_file(FINISH_TIME_FILE)

        print(f"START : {sh:02d}:{sm:02d}")
        print(f"FINISH: {fh:02d}:{fm:02d}")

        if run_at <= now + timedelta(seconds=1):
            print(
                f"\nSTART уже наступил, FINISH ещё впереди. "
                f"Запускаем полный цикл сейчас ({now:%H:%M:%S})."
            )
        else:
            _wait_until(run_at, "START")

        print("\nЗапускаю пункт 4 автоматически:")
        print("Feed -> Jobs -> Notifications -> Excel -> JSON -> GitHub")
        automate_sequence()

        print("\nПолный цикл завершён.")

        if datetime.now() < finish_at:
            _wait_until(finish_at, "FINISH")
        else:
            print("FINISH уже прошёл во время выполнения. Выключаем сейчас.")

        _shutdown_windows_now()

    except KeyboardInterrupt:
        print("\nОстановлено пользователем. ПК не выключаем.")
    except Exception as exc:
        print("\nОШИБКА AUTO MODE:")
        print(exc)
        print("При ошибке ПК автоматически НЕ выключаем.")



def main() -> None:
    parser = argparse.ArgumentParser(
        description="LinkedIn worker: ручной режим или немедленный AUTO"
    )
    parser.add_argument(
        "-a",
        "--auto",
        action="store_true",
        help="Без вопросов: профиль PC2 и полный цикл сразу",
    )
    parser.add_argument(
        "command",
        nargs="?",
        choices=["feed", "jobs", "notifications", "parse"],
        help=argparse.SUPPRESS,
    )
    parser.add_argument("folder", nargs="?", help=argparse.SUPPRESS)

    args = parser.parse_args()

    # Диспетчер запускает worker именно так:
    #   python linkedin_worker.py -a
    # Никаких таймеров и shutdown здесь нет.
    if args.auto:
        global ACTIVE_DEVICE
        ACTIVE_DEVICE = DEVICE_PRESETS["2"]

        print("=" * 66)
        print(" LinkedIn WORKER AUTO")
        print("=" * 66)
        print(f"GUI-профиль: {ACTIVE_DEVICE['name']}")
        print("Feed -> Jobs -> Notifications -> Excel -> JSON -> GitHub")

        automate_sequence()

        print("\nWORKER AUTO завершён успешно.")
        return

    # Обычный запуск файла — старый интерактивный режим.
    if args.command is None:
        choose_device()
        show_menu()
        return

    # Сохраняем и старые CLI-команды.
    choose_device()

    if args.command == "feed":
        automate("feed")
    elif args.command == "jobs":
        automate("jobs")
    elif args.command == "notifications":
        automate("notifications")
    elif args.command == "parse":
        folder = Path(args.folder) if args.folder else ROOT / "saved_pages"
        parse_folder(folder, ROOT / "config.json")


if __name__ == "__main__":
    main()
